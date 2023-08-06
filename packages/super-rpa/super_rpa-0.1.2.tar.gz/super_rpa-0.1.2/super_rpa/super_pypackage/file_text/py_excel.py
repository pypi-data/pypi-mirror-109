# -*- coding: utf-8 -*-
"""
封装操作 xlsx 文件功能：增、删、改、查
"""

from openpyxl import load_workbook


def check_excel(path):
    """
    - 显示表格所有数据，以列表形式返回。
    - 返回单元格位置是方便在修改表格的时候使用。

    :param path: 文件路径。（注意：只操作第一个 Sheet1。）
    :return: [['单元格位置', '单元格信息'], ['单元格位置', '单元格信息'], ['单元格位置', '单元格信息'], ]
    """
    wb = load_workbook(path)
    sheet = wb.active
    for row in sheet.rows:
        cell = [[row[r].coordinate, row[r].value] for r in range(len(row))]
        yield cell


def update_excel(path, index, text, sheet='Sheet1', error_log=True):
    """
    - 追加写
    :param path: 路径
    :param index: 需要插入数据的单元格位置。如：A1
    :param text: 需要插入的文本
    :param sheet:
    :param error_log:
    :return:

    使用案例：
    特点1：在表格的title（第一行）后面加上一个‘备注’ 单元格，那么下面行的单元格没有东西的话，这里会显示 None。但是一定要有一个单元格去占位置。
    for i in py_excel.check_excel('data.xlsx'):
        py_excel.update_excel('data.xlsx', i[-1][0], '测试')  # 在每行的最后一个单元格插入数据
        print(i)
    """
    ec = 1
    while True:
        try:
            xfile = load_workbook(path)
            sheet = xfile.get_sheet_by_name(sheet)
            sheet[str(index)] = str(text)
            xfile.save(path)
            print('写入成功。')
            break
        except Exception as e:
            if error_log and ec:
                ec = 0
                print(e)
                print('保存失败，请勿打开"{}"，正在重新保存。。。'.format(path))
