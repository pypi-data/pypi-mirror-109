# -*- coding: utf-8 -*-
"""
封装操作 xlsx 文件功能：增、删、改、查
"""

from openpyxl import load_workbook


def check_excel(path):
    wb = load_workbook(path)
    sheet = wb.active
    for row in sheet.rows:
        cell = [[row[r].coordinate, row[r].value] for r in range(len(row))]
        yield cell


def update_excel(path, index, text, sheet='Sheet1', error_log=True):
    ec = 1
    while True:
        try:
            xfile = load_workbook(path)
            sheet = xfile.get_sheet_by_name(sheet)
            sheet[str(index)] = str(text)
            xfile.save(path)
            print('写入成功。')
            break
        except Exception as e:
            if error_log and ec:
                ec = 0
                print(e)
                print('保存失败，请勿打开"{}"，正在重新保存。。。'.format(path))
