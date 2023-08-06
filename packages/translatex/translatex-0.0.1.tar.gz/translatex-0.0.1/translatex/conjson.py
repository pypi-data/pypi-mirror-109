import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook

class convert:

	def converting(r_excel, workbook, tojson):
		
		file_d = load_workbook(r_excel)
		sheet_name = file_d[workbook]
		table_list_j = []

		for row_d in islice(sheet_name.values, 1, sheet_name.max_row):
		    cell = OrderedDict()
		    cell['_id'] = row_d[0]
		    cell['student_name'] = row_d[1]
		    cell['class_student'] = row_d[2]
		    cell['teacher'] = row_d[3]
		    cell['subjects'] = row_d[4]
		    cell['rating'] = row_d[5]
		    table_list_j.append(cell)

		j = json.dumps(table_list_j, ensure_ascii=False)

		with open(tojson, 'w', ) as file:
			file.write(j)
		return 'THE XLSX SPREADSHEET HAS BEEN CONVERTED TO THE JSON FILE FORMAT'

